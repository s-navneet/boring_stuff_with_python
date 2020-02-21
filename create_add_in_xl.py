import openpyxl

wb = openpyxl.Workbook()  #Workbook fun create a worksheet and object refer to wrksheet
print(wb)
print(wb.get_sheet_names())  #print the sheets name
sheet=wb.get_sheet_by_name('Sheet')  #getting the sheet to work
print(str(sheet['A1'].value))   #getting the cell a1 value

#add somthing in your xl cell

sheet['A1']='name'
sheet['B1']='age'

#this above xl file is in the computer memory if youwant to stor in HD

import os
os.chdir('/home/navneet/Documents/pyxlproj')
wb.save('exapmle1.xlsx')   #save this file usingsave fun pass string as a name of file

sheet2=wb.create_sheet()   #create another sheet in same xl file
print(wb.get_sheet_names())        #

sheet2.title='My new Sheet name'  #rename the sheet name
print(wb.get_sheet_names())

wb.save('examle2.xlsx')  #create a copy of same file with add feature

wb.create_sheet(index = 0, title="new sheet1 name")  #add the new sheet in new index and give its name
wb.save('examle3.xlsx')


